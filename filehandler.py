#!/usr/bin/env pipenv-shebang
import os
import datetime
import logging

import pretty_errors
from openpyxl import Workbook, load_workbook
from openpyxl.styles import NamedStyle, Font, Alignment, numbers

import dateformat


logger = logging.getLogger('main')


def create_dirs(dirnames):
    """create new directories

    :param dirnames: directories name
    :type dirnames: list
    :return: directories path
    :rtype: list
    """
    new_dirnames = []
    for dirname in dirnames:
        dirpath = os.path.join(os.path.dirname(__file__), dirname)
        if not os.path.isdir(dirpath):
            os.mkdir(dirpath)
        new_dirnames.append(dirpath)
    return new_dirnames


def create_subdirs(base_dirpath, dirnames):
    """create new sub directories

    :param basedirpath: base directory name
    :type basedirpath: str
    :param dirnames: directories name
    :type dirnames: list
    :return: directories path
    :rtype: list
    """
    new_dirnames = []
    for dirname in dirnames:
        dirpath = os.path.join(base_dirpath, dirname)
        if not os.path.isdir(dirpath):
            os.mkdir(dirpath)
        new_dirnames.append(dirpath)
    return new_dirnames


def get_content(filepath):
    """get file content

    :param filepath: file location
    :type filepath: str
    :return: file contents, date, count, total
    :rtype: list, str, str, str
    """
    with open(filepath, 'r') as f:
        i = 1
        contents = []
        lines = f.readlines()
        max_line = len(lines)

        for line in lines:
            if i == max_line:
                r = line.split(';')
                count = r[0]
                total = int(r[1])
                total = format(total, ',')
                break
            c = line.rstrip().split('|')
            contents.append(c)
            i += 1

    filename = os.path.basename(filepath)   # 20210210_P014B.txt
    product_id = filename[9:14]             # P014B
    if product_id == 'P014B':               # PBB
        date = dateformat.format_1(c[0])
    elif product_id == 'D039P':             # PDAM
        date = dateformat.format_1(c[2])

    return contents, date, count, total


def xlsx_template_1(contents):
    """create 'Database PBB.xlsx' file

    :param contents: file contens
    :type contents: list
    :return: filename, filepath
    :rtype: str, str
    """
    filename = 'Database PBB.xlsx'
    dist_dirpath = create_dirs(['dist'])
    filepath = os.path.join(dist_dirpath[0], filename)

    # create a new file if not exist
    # or append new values if file exist
    if not os.path.isfile(filepath):
        default_font = Font(name='Calibri', size=10)

        # style template header row
        header = NamedStyle(name='header')
        header.font = Font(name='Calibri', size=10, bold=True)
        header.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

        # create a blank xlsx file
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A3:P3')

        # set column width
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].hidden = True
        ws.column_dimensions['F'].hidden = True
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].hidden = True
        ws.column_dimensions['I'].hidden = True
        ws.column_dimensions['J'].hidden = True
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].hidden = True
        ws.column_dimensions['M'].width = 15
        ws.column_dimensions['N'].width = 15
        ws.column_dimensions['O'].width = 15
        ws.column_dimensions['P'].width = 12.5
        ws.column_dimensions['Q'].width = 12.5

        # set header title
        ws['A1'] = 'PT POS INDONESIA (PERSERO)'
        ws['A1'].font = default_font
        ws['A2'] = 'KANTOR POS NGANJUK 64400'
        ws['A2'].font = default_font
        ws['A3'] = 'DATABASE TRANSAKSI PBB'
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A5'] = '#'
        ws['B5'] = 'TANGGAL'
        ws['C5'] = 'NOP'
        ws['D5'] = 'NAMA'
        ws['E5'] = 'ALAMAT'
        ws['F5'] = 'JENIS'
        ws['G5'] = 'TAHUN'
        ws['H5'] = 'STATUS'
        ws['I5'] = 'REF KODE'
        ws['J5'] = 'REF BANK'
        ws['K5'] = 'JATUH TEMPO'
        ws['L5'] = 'KODE KP'
        ws['M5'] = 'POKOK'
        ws['N5'] = 'DENDA'
        ws['O5'] = 'TAGIHAN'
        ws['P5'] = 'BEA MAINTENANCE'
        ws['Q5'] = 'BEA ADMIN'

        # set style in header column
        header_row = ws[5]
        for cell in header_row:
            cell.style = header

        append_data_1(wb, ws, contents, filepath)
    else:
        wb = load_workbook(filename=filepath)
        ws = wb.active

        append_data_1(wb, ws, contents, filepath)

    return filename, filepath


def append_data_1(wb, ws, contents, filename):
    """append new data into 'Database PBB.xlsx' file

    :param wb: workbook object
    :type wb: obj
    :param ws: worksheet object
    :type ws: obj
    :param contents: file contents
    :type contents: list
    :param filename: file name
    :type filename: str
    """
    # style template body row
    uid = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    body = NamedStyle(name=uid)
    body.font = Font(name='Calibri', size=10)

    for content in contents:
        # add increment number
        i = ws[f'A{ws.max_row}'].value
        i = 1 if i == '#' else i + 1

        # append increment number into first index
        new_content = content.copy()
        new_content.insert(0, i)
        new_content[12] = int(new_content[12])
        new_content[13] = int(new_content[13])
        new_content[14] = int(new_content[14])
        new_content[15] = int(new_content[15])
        new_content[16] = int(new_content[16])
        ws.append(new_content)

        # set style to body column
        body_row = ws[ws.max_row]
        for cell in body_row:
            cell.style = body

    # set number format in column L:Q
    col_number = ws['L:Q']
    for col in col_number:
        for cell in col[5:]:
            cell.number_format = numbers.BUILTIN_FORMATS[43]

    # save to file
    wb.save(filename)
    message = f'Append new data into {filename}'
    logger.info(message)


def xlsx_template_2(contents):
    """create 'Trx_PBB_KantorPos_{date}' file

    :param contents: file contents
    :type contents: list
    :return: file location
    :rtype: str
    """
    filename = f'Trx_PBB_KantorPos_{contents[0][0]}.xlsx'
    dist_dirpath = create_dirs(['dist'])
    base_dirpath = create_subdirs(dist_dirpath[0], ['xlsx'])
    sub_dirpath = create_subdirs(base_dirpath[0], ['pbb'])
    filepath = os.path.join(sub_dirpath[0], filename)

    # style template header row
    header = NamedStyle(name='header')
    header.font = Font(name='Calibri', size=10, bold=True)
    header.alignment = Alignment(horizontal='center', vertical='center')

    # create a blank xlsx file
    wb = Workbook()
    ws = wb.active

    # set column width
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 20

    # set header title
    ws['A1'] = 'NO'
    ws['B1'] = 'NOP'
    ws['C1'] = 'TAHUN PAJAK'
    ws['D1'] = 'NOMINAL'
    ws['E1'] = 'DENDA'
    ws['F1'] = 'TOTAL'
    ws['G1'] = 'TANGGAL TRANSAKSI'

    # set style in header column
    header_row = ws[1]
    for cell in header_row:
        cell.style = header

    # style template body row
    uid = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    body = NamedStyle(name=uid)
    body.font = Font(name='Calibri', size=10)

    # append data
    for content in contents:
        # add increment number
        i = ws[f'A{ws.max_row}'].value
        i = 1 if i == 'NO' else i + 1

        # create a new list
        new_content = []
        new_content.append(i)
        new_content.append(content[1].replace('.', ''))
        new_content.append(content[5])
        new_content.append(int(content[11]))
        new_content.append(int(content[12]))
        new_content.append(int(content[13]))
        new_content.append(dateformat.format_1(content[0]))
        ws.append(new_content)

        # set style to body column
        body_row = ws[ws.max_row]
        for cell in body_row:
            cell.style = body

    # set number format in column D:F
    col_number = ws['D:F']
    for col in col_number:
        for cell in col[1:]:
            cell.number_format = numbers.BUILTIN_FORMATS[43]

    # save to file
    wb.save(filepath)
    message = f'File {filename} has been created'
    logger.info(message)

    return filename, filepath


def xlsx_template_3(contents):
    """create 'Database PDAM.xlsx' file

    :param contents: file contents
    :type contents: list
    :return: file location
    :rtype: str
    """
    filename = 'Database PDAM.xlsx'
    dist_dirpath = create_dirs(['dist'])
    filepath = os.path.join(dist_dirpath[0], filename)

    # if file is not exist, create it
    # or append values into existing file
    if not os.path.isfile(filepath):
        default_font = Font(name='Calibri', size=10)

        # style template header row
        header = NamedStyle(name='header')
        header.font = Font(name='Calibri', size=10, bold=True)
        header.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)

        # create a blank xlsx file
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A3:G3')

        # set column width
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10

        # set header title
        ws['A1'] = 'PT POS INDONESIA (PERSERO)'
        ws['A1'].font = default_font
        ws['A2'] = 'KANTOR POS NGANJUK 64400'
        ws['A2'].font = default_font
        ws['A3'] = 'DATABASE TRANSAKSI PDAM'
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A5'] = '#'
        ws['B5'] = 'NO SAMBUNGAN'
        ws['C5'] = 'BULAN TAGIHAN'
        ws['D5'] = 'TANGGAL TRANSAKSI'
        ws['E5'] = 'TAGIHAN'
        ws['F5'] = 'ID PETUGAS'
        ws['G5'] = 'KANTOR'

        # set style in header column
        header_row = ws[5]
        for cell in header_row:
            cell.style = header

        append_data_2(wb, ws, contents, filepath)
    else:
        wb = load_workbook(filename=filepath)
        ws = wb.active

        append_data_2(wb, ws, contents, filepath)

    return filename, filepath


def append_data_2(wb, ws, contents, filename):
    """append new data into 'Database PDAM.xlsx' file

    :param wb: workbook object
    :type wb: obj
    :param ws: worksheet object
    :type ws: obj
    :param contents: file contents
    :type contents: list
    :param filename: file name
    :type filename: str
    """
    # style template body row
    uid = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    body = NamedStyle(name=uid)
    body.font = Font(name='Calibri', size=10)

    for content in contents:
        # add increment number
        i = ws[f'A{ws.max_row}'].value
        i = 1 if i == '#' else i + 1

        # append increment number into first index
        new_content = content.copy()
        new_content.insert(0, i)
        new_content[4] = int(new_content[4])
        del new_content[7]
        ws.append(new_content)

        # set style to body column
        body_row = ws[ws.max_row]
        for cell in body_row:
            cell.style = body

    # set number format in column L:Q
    col_number = ws['E']
    for cell in col_number:
        cell.number_format = numbers.BUILTIN_FORMATS[43]

    # save to file
    wb.save(filename)
    message = f'Append new data into {filename}'
    logger.info(message)
